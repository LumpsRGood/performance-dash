import streamlit as st
import pandas as pd
from datetime import datetime

import re
from datetime import datetime
from openpyxl import load_workbook

st.title("📊 Weekly Performance Dashboard")

uploaded_files = st.file_uploader(
    "Upload Sales + Turn Excel Files",
    type=["xlsx"],
    accept_multiple_files=True
)

def get_header_title(file):
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        return ws["A1"].value
    except Exception as e:
        st.warning(f"⚠️ Could not read A1 in file: {file.name} ({e})")
        return ""

def extract_dates(text):
    match = re.search(r"(\d{2}/\d{2}/\d{4}) to (\d{2}/\d{2}/\d{4})", text)
    if match:
        start = datetime.strptime(match.group(1), "%m/%d/%Y")
        end = datetime.strptime(match.group(2), "%m/%d/%Y")
        return start, end
    return None, None

parsed_files = []
for file in uploaded_files:
    header = get_header_title(file)
    start, end = extract_dates(header)
    if not start or not end:
        st.error(f"❌ Could not parse date from: {file.name}")
        continue
    if "Turn Stats" in header:
        file_type = "turn"
    elif "Sales Statistics" in header:
        file_type = "sales"
    else:
        st.warning(f"⚠️ Unrecognized file type in: {file.name}")
        continue
    parsed_files.append({"file": file, "type": file_type, "start": start, "end": end})

this_week = {}
last_week = {}

parsed_files.sort(key=lambda x: x["end"], reverse=True)
for entry in parsed_files:
    ftype = entry["type"]
    if ftype not in this_week:
        this_week[ftype] = entry
    elif ftype not in last_week:
        last_week[ftype] = entry

def label(entry):
    if entry:
        return f"{entry['file'].name} ({entry['start'].date()} → {entry['end'].date()})"
    return "Not found"

st.markdown("### 📅 File Summary")
st.write(f"**This Week Sales:** {label(this_week.get('sales'))}")
st.write(f"**This Week Turn:** {label(this_week.get('turn'))}")
st.write(f"**Last Week Sales:** {label(last_week.get('sales'))}")
st.write(f"**Last Week Turn:** {label(last_week.get('turn'))}")


st.set_page_config(page_title="Server Performance Dashboard - v1.2.39", layout="wide")

# ---------- Utility Functions ---------- #
def parse_sales(file):
    try:
        df = pd.read_excel(file, header=4)
        df.columns = df.columns.str.strip().str.lower()
        df = df[~df["location"].astype(str).str.contains("Total|Copyright|Rosnet", case=False, na=False)]
        df = df[df["employee name"].notna() & df["location"].notna()]
        df = df[df["employee name"].str.upper().str.strip() != "STAFF, OLO"]
        df["location key"] = df["location"].astype(str).str.strip()
        return df
    except Exception as e:
        st.error(f"Error reading sales file: {e}")
        return pd.DataFrame()

def parse_turn(file):
    try:
        df = pd.read_excel(file, header=4)
        df.columns = df.columns.str.strip().str.lower()
        for col in df.columns:
            if col.strip().lower() == "avg mins":
                df.rename(columns={col: "turn time"}, inplace=True)
        return df
    except Exception as e:
        st.error(f"Error reading turn file: {e}")
        return pd.DataFrame()

def merge_data(sales_df, turn_df):
    if "employee name" not in sales_df.columns or "employee name" not in turn_df.columns:
        st.error("❌ 'Employee Name' column is missing from one of the files.")
        return pd.DataFrame()
    return pd.merge(sales_df, turn_df.drop(columns=[
        col for col in turn_df.columns if col in sales_df.columns and col != "employee name"
    ]), on="employee name", how="left")

def describe_change(curr, prev, is_pct=False):
    try:
        curr = float(curr)
        prev = float(prev)
        diff = curr - prev
        if diff > 0:
            return f"Improved by {abs(diff):.2%}" if is_pct else f"Improved by {abs(diff):.2f}"
        elif diff < 0:
            return f"Declined by {abs(diff):.2%}" if is_pct else f"Declined by {abs(diff):.2f}"
        else:
            return "No Change"
    except:
        return "NEW"

def style_lw_change(val, inverse=False):
    try:
        if isinstance(val, str):
            if "NEW" in val:
                return "background-color: #f9a825; color: black; font-weight: bold; text-align: center"
            elif "Improved" in val:
                return (
                    "background-color: #1b5e20; color: white; font-weight: bold; text-align: center"
                    if not inverse else
                    "background-color: #1b5e20; color: white; font-weight: bold; text-align: center"
                )
            elif "Declined" in val:
                return (
                    "background-color: #c62828; color: white; font-weight: bold; text-align: center"
                    if not inverse else
                    "background-color: #c62828; color: white; font-weight: bold; text-align: center"
                )
        return "text-align: center"
    except:
        return "text-align: center"

def ppa_bg(val):
    try:
        v = float(val)
        if v >= 15.5:
            return "background-color: #1b5e20; color: white; text-align: center; font-weight: bold"
        else:
            return "background-color: #c62828; color: white; text-align: center; font-weight: bold"
    except:
        return "text-align: center; font-weight: bold"

def disc_pct_bg(val):
    try:
        v = float(val.strip('%')) if isinstance(val, str) else float(val)
        if v < 1.5:
            return "background-color: #1b5e20; color: white; text-align: center; font-weight: bold"
        else:
            return "background-color: #c62828; color: white; text-align: center; font-weight: bold"
    except:
        return "text-align: center; font-weight: bold"

def bev_pct_bg(val):
    try:
        v = float(val.strip('%')) if isinstance(val, str) else float(val)
        if v >= 18.5:
            return "background-color: #1b5e20; color: white; text-align: center; font-weight: bold"
        else:
            return "background-color: #c62828; color: white; text-align: center; font-weight: bold"
    except:
        return "text-align: center; font-weight: bold"

def turn_time_bg(val):
    try:
        v = float(val)
        if v <= 35:
            return "background-color: #1b5e20; color: white; text-align: center; font-weight: bold"
        else:
            return "background-color: #c62828; color: white; text-align: center; font-weight: bold"
    except:
        return "text-align: center"

def extract_first_name(full_name):
    try:
        name = full_name.replace("🏆", "").replace("🔼", "").strip()
        if "," in name:
            return name.split(",")[1].strip().split()[0]
        return name.split()[0]
    except:
        return full_name

# ---------- Highlighting Functions ---------- #
def is_top_performer(row):
    return (
        ppa_bg(row["PPA"]).startswith("background-color: #1b5e20")
        and disc_pct_bg(row["Discount %"]).startswith("background-color: #1b5e20")
        and bev_pct_bg(row["Beverage %"]).startswith("background-color: #1b5e20")
        and turn_time_bg(row["Turn Time"]).startswith("background-color: #1b5e20")
    )

def is_most_improved(row):
    try:
        return all([
            isinstance(row["+/- PPA LW"], str) and row["+/- PPA LW"].startswith("Improved"),
            isinstance(row["+/- Discount % LW"], str) and row["+/- Discount % LW"].startswith("Improved"),
            isinstance(row["+/- Beverage % LW"], str) and row["+/- Beverage % LW"].startswith("Improved"),
            isinstance(row["+/- Turn Time LW"], str) and row["+/- Turn Time LW"].startswith("Improved"),
        ])
    except:
        return False

def highlight_top_performer(row):
    if is_top_performer(row):
        return ["background-color: #2e7d32; color: white; font-weight: bold"] * len(row)
    else:
        return [""] * len(row)

def highlight_most_improved(row):
    if is_most_improved(row):
        return ["background-color: #1565c0; color: white; font-weight: bold"] * len(row)
    else:
        return [""] * len(row)

# ---------- Render Table ---------- #
def render_comparison_table(df, location):
    st.subheader(f"📍 Location: {location} Performance Comparison")
    df = df.sort_values(by="ppa", ascending=False)

    cols = ["employee name", "ppa", "+/- ppa lw", "disc %", "+/- disc % lw",
            "bev %", "+/- bev % lw", "turn time", "+/- turn lw"]
    display_df = df[cols].copy()

    display_df.rename(columns={
        "employee name": "Employee Name", "ppa": "PPA",
        "+/- ppa lw": "+/- PPA LW", "disc %": "Discount %",
        "+/- disc % lw": "+/- Discount % LW", "bev %": "Beverage %",
        "+/- bev % lw": "+/- Beverage % LW", "turn time": "Turn Time",
        "+/- turn lw": "+/- Turn Time LW"
    }, inplace=True)

    display_df["PPA"] = display_df["PPA"].map("{:.2f}".format)
    display_df["Discount %"] = display_df["Discount %"].map("{:.2%}".format)
    display_df["Beverage %"] = display_df["Beverage %"].map("{:.2%}".format)
    display_df["Turn Time"] = display_df["Turn Time"].map(lambda x: f"{x:.2f}" if pd.notnull(x) else "n/a")

    # 🏆 & 🔼 Apply labels
    top_performers = []
    most_improved = []
    for i, row in display_df.iterrows():
        name = display_df.at[i, "Employee Name"]
        badge = ""
        if is_top_performer(row):
            badge += "🏆"
            top_performers.append(extract_first_name(name))
        if is_most_improved(row):
            badge += "🔼"
            most_improved.append(extract_first_name(name))
        if badge:
            display_df.at[i, "Employee Name"] = name + " " + badge

    if top_performers:
        st.success("🏅 Top Performers: " + ", ".join(top_performers))
    if most_improved:
        st.info("🔼 Most Improved: " + ", ".join(most_improved))

    styles = display_df.style \
        .applymap(ppa_bg, subset=["PPA"]) \
        .applymap(disc_pct_bg, subset=["Discount %"]) \
        .applymap(bev_pct_bg, subset=["Beverage %"]) \
        .applymap(turn_time_bg, subset=["Turn Time"]) \
        .applymap(lambda v: style_lw_change(v, inverse=False), subset=["+/- PPA LW", "+/- Beverage % LW", "+/- Turn Time LW"]) \
        .applymap(lambda v: style_lw_change(v, inverse=True), subset=["+/- Discount % LW"]) \
        .apply(highlight_top_performer, axis=1) \
        .apply(highlight_most_improved, axis=1) \
        .set_properties(**{"text-align": "center", "vertical-align": "middle", "font-weight": "bold", "font-size": "14px"}) \
        .set_table_styles([
            {'selector': 'th', 'props': [('text-align', 'center'), ('font-weight', 'bold')]},
            {'selector': 'td', 'props': [('text-align', 'center'), ('font-weight', 'bold')]}
        ], overwrite=False)

    st.dataframe(styles, use_container_width=True, hide_index=True, height=min(800, 45 * len(display_df) + 100))

# ---------- Streamlit UI ---------- #
st.title("📊 Server Performance Dashboard – v1.2.39")

with st.expander("", expanded=True):
    st.markdown("### 📄 Upload this week's **Employee Sales Statistics**")
    this_week_file = st.file_uploader("", type="xlsx", key="tw_sales")
    st.markdown("### 📄 Upload last week's **Employee Sales Statistics**")
    last_week_file = st.file_uploader("", type="xlsx", key="lw_sales")

if this_week_file and last_week_file:
    sales_tw = parse_sales(this_week_file)
    sales_lw = parse_sales(last_week_file)

    if not sales_tw.empty and not sales_lw.empty:
        locations = sorted(sales_tw["location key"].unique())
        st.success(f"✅ Sales data uploaded! Found locations: {', '.join(locations)}")

        st.subheader("Step 2: Upload Turn Time Files")
        turn_data = {}
        for loc in locations:
            st.markdown(f"**📍 {loc}**")
            col1, col2 = st.columns(2)
            with col1:
                tw_file = st.file_uploader(f"This Week - {loc}", type="xlsx", key=f"tw_{loc}")
            with col2:
                lw_file = st.file_uploader(f"Last Week - {loc}", type="xlsx", key=f"lw_{loc}")
            turn_data[loc] = {"this_week": tw_file, "last_week": lw_file}

        if st.button("Step 3: Generate Dashboards"):
            for loc in locations:
                tw_file = turn_data[loc]["this_week"]
                lw_file = turn_data[loc]["last_week"]
                if tw_file and lw_file:
                    tw_df = parse_turn(tw_file)
                    lw_df = parse_turn(lw_file)
                    if not tw_df.empty and not lw_df.empty:
                        merged_tw = merge_data(sales_tw[sales_tw["location key"] == loc], tw_df)
                        merged_lw = merge_data(sales_lw[sales_lw["location key"] == loc], lw_df)

                        lw_index = merged_lw.set_index("employee name")

                        merged_tw["+/- ppa lw"] = merged_tw.apply(
                            lambda r: describe_change(r["ppa"], lw_index["ppa"].get(r["employee name"], None)), axis=1
                        )
                        merged_tw["+/- disc % lw"] = merged_tw.apply(
                            lambda r: describe_change(lw_index["disc %"].get(r["employee name"], None), r["disc %"], is_pct=True), axis=1
                        )
                        merged_tw["+/- bev % lw"] = merged_tw.apply(
                            lambda r: describe_change(r["bev %"], lw_index["bev %"].get(r["employee name"], None), is_pct=True), axis=1
                        )
                        merged_tw["+/- turn lw"] = merged_tw.apply(
                            lambda r: describe_change(lw_index["turn time"].get(r["employee name"], None), r["turn time"]), axis=1
                        )

                        render_comparison_table(merged_tw, loc)